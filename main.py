import tkinter as tk
import openpyxl

def execute_program():
    # Get the values entered by the user
    if start_choice_var.get() == "Ubicació actual":
        inici_location = "Ubicació actual"
    else:
        inici_location = start_entry.get()

    end_location = end_entry.get()
    lot = ruta_var.get()

    # Print the values for demonstration
    print("Començament del trajecte:", inici_location)
    print("Final del trajecte:", end_location)
    print("Lot:", lot)

    # Open the Excel file
    workbook = openpyxl.load_workbook('Dades_Municipis.xlsx')
    # Get the names of all the sheets in the workbook
    sheet_names = workbook.sheetnames

    # Ask the user to select a sheet
    if lot == '2':
        lot = int(lot)
        lot = lot-1
    elif lot == '4':
        lot = int(lot)
        lot = lot-2
    elif lot == '5':
        lot = int(lot)
        lot = lot-3
    selected_sheet_name = sheet_names[lot]  # Adjust for 0-based indexing
    # Access the selected sheet
    sheet = workbook[selected_sheet_name]
        
    # Get the dimensions of the sheet (number of rows and columns)
    max_row = sheet.max_row
    max_column = sheet.max_column

    # Print the names of all the columns in the selected sheet
    print("Available Columns:")
    for col in range(1, max_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        print(f"{col}. {column_letter}")

    # Ask the user to select a column
    selected_column_idx = int(input("Enter the index of the column you want to extract: "))
    selected_column_letter = openpyxl.utils.get_column_letter(selected_column_idx)

    # Extract the values from the selected column
    column_values = []
    for row in range(1, max_row + 1):
        cell_value = sheet[selected_column_letter + str(row)].value
        column_values.append(cell_value)

    # Create a .txt file and write the extracted column values to it
    output_file_path = f"{selected_sheet_name}_column_{selected_column_letter}.txt"
    with open(output_file_path, 'w') as txt_file:
        for value in column_values:
            txt_file.write(str(value) + '\n')

    print(f"Column values saved to '{output_file_path}'")

    # Close the workbook when done
    workbook.close()

# Create the main application window
root = tk.Tk()
root.title("Caixa d'Enginyers: BANCA COOPERATIVA")
root.geometry("450x200")

# Title label
title_label = tk.Label(root, text="BENVINGUT A BANCA D'ENGINYERS")
title_label.grid(row=0, column=0, columnspan=2, padx=10, pady=10)

# Dropdown for Start Location
start_label = tk.Label(root, text="Inici:")
start_label.grid(row=1, column=0, padx=10, pady=5, sticky="e")
start_choice_var = tk.StringVar()
start_choice_var.set("Localització actual")  # Default choice
start_choice_dropdown = tk.OptionMenu(root, start_choice_var, "Localització actual", "Tria una localització")
start_choice_dropdown.grid(row=1, column=1, padx=10, pady=5)

start_entry = tk.Entry(root)
start_entry.grid(row=1, column=2, padx=10, pady=5)
start_entry.grid_remove()  # Hide the entry initially

def toggle_start_input(*args):
    if start_choice_var.get() == "Ubicació actual":
        start_entry.grid_remove()
    elif start_choice_var.get() == "Tria una localització":
        start_entry.grid()
    else:
        start_entry.grid_remove()

start_choice_var.trace_add("write", toggle_start_input)
toggle_start_input()

# End Location Entry
end_label = tk.Label(root, text="Final del trajecte:")
end_label.grid(row=2, column=0, padx=10, pady=5, sticky="e")
end_entry = tk.Entry(root)
end_entry.grid(row=2, column=1, padx=10, pady=5)

# Ruta Entry
ruta_label = tk.Label(root, text="Lot:")
ruta_label.grid(row=3, column=0, padx=10, pady=5, sticky="e")
ruta_var = tk.StringVar()
ruta_dropdown = tk.OptionMenu(root, ruta_var, "2", "4", "5")
ruta_dropdown.grid(row=3, column=1, padx=10, pady=5)

# Button to execute the program
execute_button = tk.Button(root, text="Execute", command=execute_program)
execute_button.grid(row=4, column=0, columnspan=2, pady=10)

# Run the main event loop
root.mainloop()
