import openpyxl

# Open the Excel file
workbook = openpyxl.load_workbook('Dades_Municipis.xlsx')

# Get the names of all the sheets in the workbook
sheet_names = workbook.sheetnames

# Print the names of all the sheets
print("Available Sheets:")
for idx, name in enumerate(sheet_names, start=1):
    print(f"{idx}. {name}")

# Ask the user to select a sheet
selected_sheet_idx = int(input("Enter the index of the sheet you want to access: "))
selected_sheet_name = sheet_names[selected_sheet_idx - 1]  # Adjust for 0-based indexing

# Access the selected sheet
sheet = workbook[selected_sheet_name]

# Get the dimensions of the sheet (number of rows and columns)
max_row = sheet.max_row
max_column = sheet.max_column

# Print the names of all the columns in the selected sheet
print("Available Columns:")
for col in range(1, max_column + 1):
    column_letter = openpyxl.utils.get_column_letter(col)
    print(f"{col}. {column_letter}")

# Ask the user to select a column
selected_column_idx = int(input("Enter the index of the column you want to extract: "))
selected_column_letter = openpyxl.utils.get_column_letter(selected_column_idx)

# Extract the values from the selected column
column_values = []
for row in range(1, max_row + 1):
    cell_value = sheet[selected_column_letter + str(row)].value
    column_values.append(cell_value)

# Create a .txt file and write the extracted column values to it
output_file_path = f"{selected_sheet_name}_column_{selected_column_letter}.txt"
with open(output_file_path, 'w') as txt_file:
    for value in column_values:
        txt_file.write(str(value) + '\n')

print(f"Column values saved to '{output_file_path}'")

# Close the workbook when done
workbook.close()
